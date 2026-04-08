"""FastAPI 路由测试"""
import pytest
import io
from pathlib import Path
from fastapi.testclient import TestClient
from db.models import Base, engine, init_db
from api.server import app

client = TestClient(app)

TEST_DIR = Path(__file__).parent / "test_data"
TEST_DIR.mkdir(exist_ok=True)


@pytest.fixture(autouse=True)
def setup_db():
    """每个测试前重建数据库"""
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    yield
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)


def _make_txt_file(content: str = "测试文档内容", filename: str = "test.txt"):
    """创建一个内存中的txt文件用于上传"""
    return ("file", (filename, io.BytesIO(content.encode("utf-8")), "text/plain"))


class TestHealthCheck:
    def test_health(self):
        resp = client.get("/api/health")
        assert resp.status_code == 200
        assert resp.json()["status"] == "ok"


class TestDocumentAPI:
    def test_upload_document(self):
        resp = client.post("/api/documents/upload", files=[_make_txt_file()])
        assert resp.status_code == 200
        data = resp.json()
        assert "id" in data
        assert data["filename"] == "test.txt"

    def test_upload_unsupported_format(self):
        file = ("file", ("test.xyz", io.BytesIO(b"data"), "application/octet-stream"))
        resp = client.post("/api/documents/upload", files=[file])
        assert resp.status_code == 400

    def test_list_documents(self):
        # 先上传一个文档
        client.post("/api/documents/upload", files=[_make_txt_file()])
        resp = client.get("/api/documents")
        assert resp.status_code == 200
        docs = resp.json()
        assert len(docs) >= 1
        assert docs[0]["filename"] == "test.txt"

    def test_parse_document(self):
        # 上传
        upload_resp = client.post("/api/documents/upload", files=[_make_txt_file("这是解析测试内容")])
        doc_id = upload_resp.json()["id"]
        # 解析
        resp = client.post(f"/api/documents/parse/{doc_id}")
        assert resp.status_code == 200
        assert resp.json()["text_length"] > 0

    def test_parse_nonexistent(self):
        resp = client.post("/api/documents/parse/99999")
        assert resp.status_code == 404

    def test_delete_document(self):
        upload_resp = client.post("/api/documents/upload", files=[_make_txt_file()])
        doc_id = upload_resp.json()["id"]
        resp = client.delete(f"/api/documents/{doc_id}")
        assert resp.status_code == 200
        # 确认已删除
        docs = client.get("/api/documents").json()
        assert all(d["id"] != doc_id for d in docs)

    def test_delete_nonexistent(self):
        resp = client.delete("/api/documents/99999")
        assert resp.status_code == 404


class TestEntityAPI:
    def test_list_entities_empty(self):
        resp = client.get("/api/entities")
        assert resp.status_code == 200
        assert resp.json() == []

    def test_list_entities_with_data(self):
        from db.database import DocumentDAO, EntityDAO
        doc = DocumentDAO.create("test.txt", "txt", "/tmp/test.txt")
        EntityDAO.create_batch(doc.id, [
            {"type": "person", "value": "张三", "context": "负责人张三", "confidence": 0.9},
            {"type": "phone", "value": "13800138000", "context": "电话13800138000", "confidence": 0.95},
        ])
        resp = client.get("/api/entities")
        assert resp.status_code == 200
        entities = resp.json()
        assert len(entities) == 2

    def test_search_entities_by_keyword(self):
        from db.database import DocumentDAO, EntityDAO
        doc = DocumentDAO.create("test.txt", "txt", "/tmp/test.txt")
        EntityDAO.create_batch(doc.id, [
            {"type": "person", "value": "李四", "confidence": 0.9},
            {"type": "person", "value": "王五", "confidence": 0.85},
        ])
        resp = client.get("/api/entities", params={"keyword": "李四"})
        assert resp.status_code == 200
        assert len(resp.json()) == 1
        assert resp.json()[0]["value"] == "李四"

    def test_filter_entities_by_doc(self):
        from db.database import DocumentDAO, EntityDAO
        doc1 = DocumentDAO.create("a.txt", "txt", "/tmp/a.txt")
        doc2 = DocumentDAO.create("b.txt", "txt", "/tmp/b.txt")
        EntityDAO.create_batch(doc1.id, [{"type": "person", "value": "A", "confidence": 0.9}])
        EntityDAO.create_batch(doc2.id, [{"type": "person", "value": "B", "confidence": 0.9}])
        resp = client.get("/api/entities", params={"doc_id": doc1.id})
        assert resp.status_code == 200
        assert len(resp.json()) == 1
        assert resp.json()[0]["value"] == "A"

    def test_export_entities_csv(self):
        from db.database import DocumentDAO, EntityDAO
        doc = DocumentDAO.create("test.txt", "txt", "/tmp/test.txt")
        EntityDAO.create_batch(doc.id, [
            {"type": "person", "value": "张三", "context": "负责人张三", "confidence": 0.9},
        ])
        resp = client.get("/api/entities/export", params={"fmt": "csv"})
        assert resp.status_code == 200
        assert "text/csv" in resp.headers["content-type"]
        text = resp.content.decode("utf-8-sig")
        assert "type,value,context,confidence" in text
        assert "person" in text
        assert "张三" in text

    def test_export_entities_xlsx(self):
        from openpyxl import load_workbook
        from db.database import DocumentDAO, EntityDAO
        doc = DocumentDAO.create("test.txt", "txt", "/tmp/test.txt")
        EntityDAO.create_batch(doc.id, [
            {"type": "phone", "value": "13800138000", "context": "电话13800138000", "confidence": 0.95},
        ])
        resp = client.get("/api/entities/export", params={"fmt": "xlsx"})
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        workbook = load_workbook(io.BytesIO(resp.content))
        sheet = workbook.active
        assert sheet["A1"].value == "id"
        assert sheet["B2"].value == "phone"
        assert sheet["C2"].value == "13800138000"


class TestTemplateAPI:
    def _make_xlsx(self):
        """创建一个简单的xlsx模板文件"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "电话", "邮箱"])
        ws.append(["", "", ""])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return ("file", ("template.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))

    def test_upload_template(self):
        resp = client.post("/api/templates/upload", files=[self._make_xlsx()])
        assert resp.status_code == 200
        data = resp.json()
        assert "id" in data
        assert "fields" in data
        assert len(data["fields"]) > 0


class TestArticleAPI:
    def test_list_articles_empty(self):
        resp = client.get("/api/articles")
        assert resp.status_code == 200
        assert resp.json() == []

    def test_list_articles_with_data(self):
        from db.database import CrawledArticleDAO
        CrawledArticleDAO.create("测试新闻", "记者", "新华网", "http://example.com",
                                  "2026-03-18", "新闻内容", "热点")
        resp = client.get("/api/articles")
        assert resp.status_code == 200
        assert len(resp.json()) == 1
        assert resp.json()[0]["title"] == "测试新闻"

    def test_get_article_detail(self):
        from db.database import CrawledArticleDAO
        article = CrawledArticleDAO.create("详情测试", "作者", "来源", "http://test.com",
                                            "2026-03-20", "详细内容", "科技")
        resp = client.get(f"/api/articles/{article.id}")
        assert resp.status_code == 200
        data = resp.json()
        assert data["title"] == "详情测试"
        assert data["content"] == "详细内容"

    def test_get_article_nonexistent(self):
        resp = client.get("/api/articles/99999")
        assert resp.status_code == 404


class TestStatisticsAPI:
    def test_statistics_empty(self):
        resp = client.get("/api/statistics")
        assert resp.status_code == 200
        data = resp.json()
        assert data["documents"] == 0
        assert data["entities"] == 0
        assert data["templates"] == 0
        assert data["articles"] == 0

    def test_statistics_with_data(self):
        from db.database import DocumentDAO, EntityDAO
        doc = DocumentDAO.create("test.txt", "txt", "/tmp/test.txt")
        EntityDAO.create_batch(doc.id, [
            {"type": "person", "value": "张三", "confidence": 0.9},
            {"type": "organization", "value": "测试公司", "confidence": 0.85},
        ])
        resp = client.get("/api/statistics")
        data = resp.json()
        assert data["documents"] == 1
        assert data["entities"] == 2
        assert "person" in data["entity_types"]
        assert data["entity_types"]["person"] == 1
